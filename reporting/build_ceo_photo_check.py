"""
CEO photo checking utilities.
Builds validation lists from the photo pool and CSV data, recomputes scores,
and can generate Excel reports with embedded photos.
"""

import argparse
from pathlib import Path
import json
import random
from typing import List, Dict, Any, Optional
import pandas as pd

try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def calculate_score(image_path: Path) -> int:
    """
    Heuristic scoring for a photo.
    4: Single-CEO photo, large (width or height > 500px).
    3: Single-CEO photo, smaller (width or height <= 500px).
    2: Group photo (filename ends with _1).
    1: No photo (file missing).
    """
    try:
        if not image_path.exists():
            return 1
        

        if image_path.stem.endswith('_1'):
            return 2
        
        if not PIL_AVAILABLE:
            return 3
        
        img = PILImage.open(image_path)
        width, height = img.size
        


        if width > 500 or height > 500:
            return 4
        else:
            return 3
    except Exception:
        return 3


def generate_excel_report(csv_path: Path, photo_pool_path: Path, output_excel_path: Path, 
                          cell_width: float, cell_height: float, insert_photos: bool = False) -> None:
    """
    Create an Excel report including CSV columns, score columns, and photo previews.
    """
    if not OPENPYXL_AVAILABLE:
        print("[ERROR] openpyxl not installed; cannot generate Excel file")
        return

    if not csv_path.exists():
        print(f"[ERROR] CSV not found: {csv_path}")
        return


    print(f"Photo pool path: {photo_pool_path.absolute()}")
    if not photo_pool_path.exists():
        print(f"[WARNING] Photo pool directory not found: {photo_pool_path}")
    else:
        png_count = len(list(photo_pool_path.glob('*.png')))
        print(f"Photo pool contains {png_count} PNG files")
        if png_count > 0:
            print("Photo samples:")
            for p in list(photo_pool_path.glob('*.png'))[:3]:
                print(f"  - {p.name}")
        else:
            print("[WARNING] No PNG files found in photo pool")

    print(f"Reading CSV: {csv_path}")
    try:
        df = pd.read_csv(csv_path)
        

        df['cik_numeric'] = pd.to_numeric(df['cik'], errors='coerce')
        

        mask = df['cik_numeric'].apply(lambda x: x != float('inf') and x != float('-inf'))
        df.loc[~mask, 'cik_numeric'] = float('nan')
        

        df['cik_numeric'] = df['cik_numeric'].fillna(0)
        

        df['cik10_str'] = df['cik_numeric'].astype(int).astype(str).str.zfill(10)
        
    except Exception as e:
        print(f"[ERROR] Failed to read CSV: {e}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "CEO Photo Check"
    

    original_headers = list(df.columns)

    if 'cik10_str' in original_headers:
        original_headers.remove('cik10_str')
        




    new_headers = ['Score_Filled', 'Score_NoFill', 'Fill_Flag', 'Photo']
    all_headers = original_headers + new_headers
    

    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = cell.font.copy(bold=True)
    

    score_filled_col_idx = len(original_headers) + 1
    score_nofill_col_idx = len(original_headers) + 2
    fill_flag_col_idx = len(original_headers) + 3
    photo_col_idx = len(original_headers) + 4
    



    ws.column_dimensions[get_column_letter(photo_col_idx)].width = cell_width / 7
    
    print(f"Processing {len(df)} rows...")
    
    inserted_count = 0
    
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2
        

        for col_idx, header in enumerate(original_headers, start=1):
            val = row[header]

            if pd.isna(val):
                val = ""
            ws.cell(row=excel_row, column=col_idx).value = val
            

        cik10 = row['cik10_str']
        

        if 'fyear' in row:
            year_val = row['fyear']
        elif 'year' in row:
            year_val = row['year']
        else:
            year_val = 0
            
        try:
            year = int(year_val)
        except (ValueError, TypeError):
            year = 0
        




        photo_std_name = f"{cik10}_{year}.png"
        photo_fill_name = f"{cik10}_{year}_fill.png"
        photo_group_name = f"{cik10}_{year}_1.png"
        
        photo_std = photo_pool_path / photo_std_name
        photo_fill = photo_pool_path / photo_fill_name
        photo_group = photo_pool_path / photo_group_name
        

        is_filled = False
        
        if photo_std.exists():
            photo_path = photo_std
            photo_filename = photo_std_name
        elif photo_fill.exists():
            photo_path = photo_fill
            photo_filename = photo_fill_name
            is_filled = True
        elif photo_group.exists():
            photo_path = photo_group
            photo_filename = photo_group_name
        else:
            photo_path = photo_std
            photo_filename = photo_std_name
        

        score = calculate_score(photo_path)
        


        ws.cell(row=excel_row, column=score_filled_col_idx).value = score

        ws.cell(row=excel_row, column=score_nofill_col_idx).value = 1 if is_filled else score

        ws.cell(row=excel_row, column=fill_flag_col_idx).value = 1 if is_filled else 0

        ws.cell(row=excel_row, column=photo_col_idx).value = 1 if photo_path.exists() else ""
        


        ws.row_dimensions[excel_row].height = cell_height * 0.75
        

        if insert_photos and photo_path.exists():
            try:
                img = OpenpyxlImage(photo_path)
                


                img_w, img_h = img.width, img.height
                
                scale_w = cell_width / img_w
                scale_h = cell_height / img_h
                scale = min(scale_w, scale_h)
                

                scale *= 0.9
                
                img.width = int(img_w * scale)
                img.height = int(img_h * scale)
                

                cell_address = f"{get_column_letter(photo_col_idx)}{excel_row}"
                ws.add_image(img, cell_address)
                inserted_count += 1
                
            except Exception as e:
                print(f"[WARNING] Unable to insert image {photo_filename}: {e}")
        
        if row_idx % 100 == 0:
            print(f"  Processed {row_idx}/{len(df)} rows...")


    output_excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_excel_path)
    print(f"[OK] Excel saved: {output_excel_path}")
    print(f"  Rows processed: {len(df)}")
    if insert_photos:
        print(f"  Photos inserted: {inserted_count}")
    else:
        print("  Photos not inserted (enable --insert-photos to embed)")


def main():
    parser = argparse.ArgumentParser(description="Generate CEO photo check Excel report")
    parser.add_argument(
        "--photo-pool",
        type=Path,
        default=Path("D:/ceo_photo_pool"),
        help="Photo pool directory path",
    )
    parser.add_argument(
        "--csv",
        type=Path,
        default=Path("scrapephoto") / "execucomp_company_identifiers_annual_2000_2025_with_ceo_FINAL.csv",
        help="Path to CEO info CSV file",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("scrapephoto") / "ceo_photo_check.xlsx",
        help="Output Excel file path",
    )
    parser.add_argument(
        "--insert-photos",
        action="store_true",
        help="Embed photos in Excel (off by default)",
    )
    parser.add_argument(
        "--cell-width",
        type=float,
        default=80.0,
        help="Photo cell width in pixels (default: 80)",
    )
    parser.add_argument(
        "--cell-height",
        type=float,
        default=80.0,
        help="Photo cell height in pixels (default: 80)",
    )
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("CEO Photo Check report generator")
    print("=" * 60)
    
    generate_excel_report(
        args.csv,
        args.photo_pool,
        args.output,
        args.cell_width,
        args.cell_height,
        args.insert_photos
    )

if __name__ == "__main__":
    main()

# python scrapephoto/build_ceo_photo_check.py
# python scrapephoto/build_ceo_photo_check.py --insert-photos
